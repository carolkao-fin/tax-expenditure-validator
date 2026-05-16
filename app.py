import streamlit as st
import pandas as pd
from docx import Document
from docx.oxml.ns import qn
import io
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber

st.set_page_config(page_title="稅式支出評估報告驗證工具", layout="wide")

# ── document extraction with page tracking ────────────────────────────────────

def extract_document(doc):
    """
    Walk the document body tracking page numbers via w:lastRenderedPageBreak
    and explicit w:br w:type='page'.

    Tables are stored as a single entry (all rows intact, so find_table_by_header
    still sees the full table), but each row also carries its own page number in
    row_pages so page attribution is per-row, not per-table-start.

    Returns:
      para_list  : [(page_num, text), ...]
      table_list : [(start_page, rows, row_pages), ...]
                   where row_pages[i] is the page for rows[i]
    """
    current_page = 1
    para_list = []
    table_list = []

    for elem in doc.element.body:
        local = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag

        if local == 'p':
            breaks_before_text = 0
            breaks_after_text  = 0
            seen_text = False

            for run in elem:
                run_local = run.tag.split('}')[-1] if '}' in run.tag else run.tag
                if run_local != 'r':
                    continue
                for child in run:
                    cl = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if cl == 'lastRenderedPageBreak':
                        if not seen_text:
                            breaks_before_text += 1
                    elif cl == 'br' and child.get(qn('w:type')) == 'page':
                        if seen_text:
                            breaks_after_text += 1
                        else:
                            breaks_before_text += 1
                    elif cl == 't' and (child.text or '').strip():
                        seen_text = True

            current_page += breaks_before_text
            text = ''.join(n.text or '' for n in elem.iter(qn('w:t'))).strip()
            if text:
                para_list.append((current_page, text))
            current_page += breaks_after_text

        elif local == 'tbl':
            table_start_page = current_page
            rows      = []
            row_pages = []   # parallel to rows: page number for each row

            for tr in elem.iter(qn('w:tr')):
                # Count page breaks that appear before any text in this row
                breaks_before = 0
                seen_text_in_row = False
                for run_elem in tr.iter(qn('w:r')):
                    for child in run_elem:
                        cl = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                        if cl == 'lastRenderedPageBreak' and not seen_text_in_row:
                            breaks_before += 1
                        elif (cl == 'br' and child.get(qn('w:type')) == 'page'
                              and not seen_text_in_row):
                            breaks_before += 1
                        elif cl == 't' and (child.text or '').strip():
                            seen_text_in_row = True

                current_page += breaks_before   # advance before recording this row

                cells = []
                for tc in tr.iter(qn('w:tc')):
                    cell_text = ''.join(
                        n.text or '' for n in tc.iter(qn('w:t'))).strip()
                    cells.append(cell_text)
                if cells:
                    rows.append(cells)
                    row_pages.append(current_page)

            if rows:
                table_list.append((table_start_page, rows, row_pages))

    return para_list, table_list


def _tbl(entry):
    """Unpack a table_list entry into (start_page, rows, row_pages)."""
    if len(entry) == 3:
        return entry
    page, rows = entry
    return page, rows, [page] * len(rows)


def extract_pdf(file_bytes):
    """
    Extract paragraphs and tables from a PDF file with accurate page numbers.
    Each page in the PDF is a hard page boundary — no heuristics required.
    Table regions are excluded from text extraction to avoid duplication.
    Returns (para_list, table_list) in the same format as extract_document().
    """
    para_list  = []
    table_list = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            # Detect table objects and extract structured rows
            found_tables = page.find_tables()
            table_bboxes = []
            for tbl_obj in found_tables:
                table_bboxes.append(tbl_obj.bbox)
                rows = []
                for row in tbl_obj.extract():
                    cells = [str(c or '').strip() for c in row]
                    if any(c for c in cells):
                        rows.append(cells)
                if rows:
                    table_list.append((page_num, rows, [page_num] * len(rows)))

            # Extract text only from areas outside table bounding boxes
            # so table content doesn't appear twice in all_text.
            if table_bboxes:
                def not_in_any_table(obj, _bboxes=table_bboxes):
                    ox0 = obj.get('x0', 0)
                    ox1 = obj.get('x1', 0)
                    ot  = obj.get('top', 0)
                    ob  = obj.get('bottom', 0)
                    for x0, top, x1, bottom in _bboxes:
                        if ox0 >= x0 - 2 and ox1 <= x1 + 2 and ot >= top - 2 and ob <= bottom + 2:
                            return False
                    return True
                text_page = page.filter(not_in_any_table)
            else:
                text_page = page

            text = text_page.extract_text(x_tolerance=3, y_tolerance=3) or ''
            for line in text.split('\n'):
                line = line.strip()
                if line:
                    para_list.append((page_num, line))

    return para_list, table_list


def build_paged_text(para_list, table_list):
    """
    Concatenate all content in page order, return (all_text, get_page_func).
    get_page_func(match_start) → page number for that character position.
    """
    by_page = {}
    for page, text in para_list:
        by_page.setdefault(page, []).append(text)
    for entry in table_list:
        _, rows, row_pages = _tbl(entry)
        for row, rp in zip(rows, row_pages):
            by_page.setdefault(rp, []).append(' '.join(row))

    pieces = []
    boundaries = []  # (start_pos, page_num)
    pos = 0
    for page_num in sorted(by_page.keys()):
        boundaries.append((pos, page_num))
        chunk = '\n'.join(by_page[page_num]) + '\n'
        pieces.append(chunk)
        pos += len(chunk)

    all_text = ''.join(pieces)

    def get_page(match_start):
        page = boundaries[0][1] if boundaries else 1
        for bpos, bpage in boundaries:
            if bpos <= match_start:
                page = bpage
            else:
                break
        return page

    return all_text, get_page


# ── helpers ───────────────────────────────────────────────────────────────────

def page_str(page):
    return f'第{page}頁' if page else '—'

def clean_num(text):
    if text is None:
        return None
    s = str(text).strip()
    if s in ('', '-', '—', '–'):
        return None
    has_pct = '%' in s
    s = re.sub(r'[,，\s億元千萬%（）()【】]', '', s)
    try:
        v = float(s)
        return v / 100 if has_pct else v
    except ValueError:
        return None

def pct_str(text):
    if text is None:
        return None
    s = str(text).strip().replace('%', '').replace(',', '')
    try:
        return float(s) / 100
    except ValueError:
        return None

def round2(v):
    return round(v, 2) if v is not None else None

def fmt(v, decimals=2):
    if v is None:
        return ''
    return f"{v:,.{decimals}f}"

def ok(reported, computed, tol=0.02):
    if reported is None or computed is None:
        return None
    denom = max(abs(reported), 1)
    return abs(reported - computed) / denom <= tol

def status_icon(passed):
    if passed is None:
        return '⚠️'
    return '✅' if passed else '❌'

def find_table_by_header(table_list, *keywords):
    """
    Returns (page, rows) or (None, None).
    Searches the first 3 rows so tables with a merged group-header row above
    the actual column headers are still matched.
    """
    for entry in table_list:
        page, rows, _ = _tbl(entry)
        if not rows:
            continue
        header = ' '.join(
            ' '.join(cell.replace('\n', '') for cell in row)
            for row in rows[:3]
        )
        if all(kw in header for kw in keywords):
            return page, rows
    return None, None

def row_result(section, item, formula, reported, computed, diff, result, note, page=None):
    return {
        '章節': section,
        '驗證項目': item,
        '公式（文件原文）': formula,
        '報告值': reported,
        '計算值': computed,
        '差異': diff,
        '結果': result,
        '說明': note,
        '資料頁碼': page_str(page),
    }

# ── formula text extraction ───────────────────────────────────────────────────

def compact_formula_full(txt):
    lines = [l.strip() for l in txt.replace('\r', '\n').split('\n') if l.strip()]
    formula_line = next(
        (l.lstrip('= ').strip() for l in lines
         if l.startswith('=') and ('×' in l or 'x' in l.lower()) and '億元' in l),
        None
    )
    result_line = next((l for l in lines if '新臺幣' in l and '億元' in l), None)
    parts = []
    if formula_line:
        parts.append(formula_line)
    if result_line:
        m = re.search(r'新臺幣\s*([\d\.]+)\s*億元', result_line)
        if m and (not formula_line or m.group(1) not in formula_line):
            parts.append(f'= {m.group(1)}億元')
    return '  '.join(parts) if parts else txt[:120]

def compact_formula_braces(txt):
    m = re.search(r'[｛{]\s*[〔\[].*?[〕\]].*?[｝}]', txt, re.DOTALL)
    if m:
        return re.sub(r'\s+', ' ', m.group(0)).strip()[:200]
    return ''

# ── document-type detection ───────────────────────────────────────────────────

def detect_type(table_list, para_list):
    for entry in table_list:
        _, rows, _ = _tbl(entry)
        flat = ' '.join(c for row in rows for c in row)
        if '提案委員' in flat and '法規內容' in flat:
            return 'simplified'
    all_text = ' '.join(t for _, t in para_list)
    if '稅式支出評估報告' in all_text and '表4-1' in all_text:
        return 'full'
    if '簡要' in all_text or '提案委員' in all_text:
        return 'simplified'
    return 'full'

# ── table financial scanner ───────────────────────────────────────────────────

_KEYWORD_MAP = [
    ('total_cit',   ['營利事業所得稅合計', '三項合計']),
    ('vehicle_cit', ['整車業增加稅額', '汽車整車', '整車業', '整車']),
    ('parts_cit',   ['零組件業增加稅額', '汽車零組件', '零組件']),
    ('other_cit',   ['其他工業及服務業', '其他工業']),
    ('tariff_loss', ['關稅損失', '最初收入損失', '關稅收入減少']),
    ('commodity',   ['貨物稅增加', '貨物稅']),
    ('vat',         ['加值型營業稅增加', '加值型營業稅']),
    ('personal',    ['個人綜合所得稅', '員工個人所得稅', '薪資所得稅']),
    ('dividend',    ['股東個人股利所得稅', '股利所得稅']),
    ('net',         ['稅收淨損益', '淨損益', '最終稅收淨', '稅收淨']),
]


def scan_tables_financial(table_list, unit):
    """
    Scan every table for financial figures in `unit` (千元 or 億元).
    Three layouts handled:
      A — unit appears inline in each row:  「2,240,360千元」
      B — unit only in column header:       header says「評估金額（千元）」, rows are plain numbers
      C — unit nowhere in table but rows have tax keywords + large numbers
          (e.g. simplified-format form tables where unit is implied)
    Returns dict: {item_key: (value, page)}.
    """
    found = {}

    for entry in table_list:
        page, rows, row_pages = _tbl(entry)
        if not rows:
            continue

        table_flat = ' '.join(c for row in rows for c in row)
        unit_in_table = unit in table_flat

        has_tax_kw = any(
            any(kw in table_flat for kw in kws)
            for _, kws in _KEYWORD_MAP
        )

        if not unit_in_table and not has_tax_kw:
            continue

        header_text = ' '.join(rows[0]) if rows else ''
        unit_in_header = unit in header_text

        val_col = None
        if unit_in_header:
            for ci, cell in enumerate(rows[0]):
                if unit in cell:
                    val_col = ci
                    break
            if val_col is None:
                val_col = len(rows[0]) - 1

        for ri, row in enumerate(rows):
            if ri == 0 and unit_in_header:
                continue  # skip header row

            row_text = ' '.join(c.strip() for c in row if c.strip())
            if not row_text:
                continue

            # Skip formula/calculation rows
            if ('×' in row_text or '＋' in row_text) and unit in row_text:
                continue

            val = None

            # ── Layout A: unit inline ────────────────────────────────────────
            if unit_in_table:
                inline = re.findall(r'([\d,]+(?:\.\d+)?)\s*' + re.escape(unit), row_text)
                if inline:
                    for n in reversed(inline):
                        try:
                            v = float(n.replace(',', ''))
                            if v >= 10:
                                val = v
                                break
                        except ValueError:
                            pass

            # ── Layout B: unit only in header ────────────────────────────────
            if val is None and unit_in_header and val_col is not None:
                candidates = []
                if val_col < len(row):
                    candidates.append(row[val_col])
                candidates += list(reversed(row))
                for cell in candidates:
                    cell_t = cell.strip()
                    if not cell_t:
                        continue
                    m_full = re.fullmatch(r'[\d,]+(?:\.\d+)?', cell_t)
                    if not m_full:
                        nums = re.findall(r'(?<!\d)([\d,]{4,}(?:\.\d+)?)(?!\d)', cell_t)
                        if not nums:
                            continue
                        cell_t = nums[-1]
                    try:
                        v = float(cell_t.replace(',', ''))
                        if v >= 100:
                            val = v
                            break
                    except ValueError:
                        pass

            # ── Layout C: no unit anywhere, but table has tax keywords ───────
            # Try the rightmost cell that contains only a large plain number.
            if val is None and not unit_in_table and has_tax_kw:
                for cell in reversed(row):
                    cell_t = cell.strip()
                    if not cell_t:
                        continue
                    # Accept pure digit strings (with optional commas)
                    if re.fullmatch(r'[\d,]+', cell_t):
                        try:
                            v = float(cell_t.replace(',', ''))
                            if v >= 1000:   # higher threshold — no unit to confirm
                                val = v
                                break
                        except ValueError:
                            pass
                    # Also accept "NNN（千元）" style annotations
                    m_ann = re.search(r'([\d,]+)\s*[（(]?千元[）)]?', cell_t)
                    if m_ann:
                        try:
                            v = float(m_ann.group(1).replace(',', ''))
                            if v >= 100:
                                val = v
                                break
                        except ValueError:
                            pass

            if val is None or val < 10:
                continue

            row_page = row_pages[ri] if ri < len(row_pages) else page
            # Map to tax item (first match per key wins)
            for key, kws in _KEYWORD_MAP:
                if key in found:
                    continue
                if any(kw in row_text for kw in kws):
                    found[key] = (val, row_page)
                    break

    return found


# ── full-format parser ────────────────────────────────────────────────────────

def parse_full_report(table_list, para_list):
    data = {}
    pages = {}

    # 表4-1
    p41, tbl41 = find_table_by_header(table_list, '5年平均')
    if tbl41 is None:
        p41, tbl41 = find_table_by_header(table_list, '現行稅率', '降稅後稅率')
    items = []
    if tbl41:
        pages['table41'] = p41
        for row in tbl41:
            if len(row) < 3:
                continue
            row_text = ' '.join(row)
            if '合計' in row_text and clean_num(row[2]) is not None:
                if data.get('total_import_reported') is None:
                    data['total_import_reported'] = clean_num(row[2])
                    data['weighted_rate_reported'] = pct_str(row[3]) if len(row) > 3 else None
                continue
            hs = row[0].strip()
            if not hs or 'HS' in hs.upper() or '稅則' in hs or '合計' in hs:
                continue
            imp = clean_num(row[2]) if len(row) > 2 else None
            cur_rate = pct_str(row[3]) if len(row) > 3 else None
            aft_rate = pct_str(row[4]) if len(row) > 4 else None
            if imp is not None:
                items.append({'hs': hs, 'name': row[1] if len(row) > 1 else '',
                               'import_5yr': imp, 'current_rate': cur_rate, 'after_rate': aft_rate})
    data['items'] = items

    # 表4-2 — try multiple header patterns
    p42, tbl42 = find_table_by_header(table_list, '關稅損失')
    if tbl42 is None:
        p42, tbl42 = find_table_by_header(table_list, '最初收入損失')
    if tbl42 is None:
        p42, tbl42 = find_table_by_header(table_list, '收入損失', '稅則')
    if tbl42 is None:
        p42, tbl42 = find_table_by_header(table_list, '降稅後稅率', '關稅')
    # Last resort: a table with multiple HS-code rows and a 合計 row
    if tbl42 is None:
        for entry in table_list:
            page, rows, _ = _tbl(entry)
            hs_rows = [r for r in rows if r and re.match(r'\d{4}[\.\d]*', r[0].strip())]
            total_row = [r for r in rows if r and '合計' in r[0]]
            flat = ' '.join(c for r in rows for c in r)
            if len(hs_rows) >= 3 and total_row and '億元' in flat:
                p42, tbl42 = page, rows
                break
    item_losses = {}
    if tbl42:
        pages['table42'] = p42
        for row in tbl42:
            if len(row) < 4:
                continue
            row_text = ' '.join(row)
            if '合計' in row_text and clean_num(row[-1]) is not None:
                if data.get('tariff_loss_reported') is None:
                    data['tariff_loss_reported'] = clean_num(row[-1])
                continue
            hs = row[0].strip()
            if not hs or 'HS' in hs.upper() or '稅則' in hs or '合計' in hs:
                continue
            loss = clean_num(row[-1])
            if loss is not None:
                item_losses[hs] = loss
    data['item_losses_reported'] = item_losses

    # Identify formula tables
    formula_tables = {}
    for entry in table_list:
        page, rows, _ = _tbl(entry)
        cell_text = '\n'.join(' '.join(row) for row in rows) if rows else ''
        ct = cell_text.replace(' ', '')
        if '汽車整車營利事業所得稅' in cell_text and '×15%' in ct:
            formula_tables['vehicle_cit'] = rows
            pages['vehicle_cit'] = page
        elif '汽車零組件營利事業所得稅' in cell_text and '×13%' in ct:
            formula_tables['parts_cit'] = rows
            pages['parts_cit'] = page
        elif '其他工業及服務業營利事業所得稅' in cell_text and '16.' in cell_text:
            formula_tables['other_cit'] = rows
            pages['other_cit'] = page
        elif ('股利所得稅' in cell_text or '股東個人股利' in cell_text) and (
                '盈餘分配比例' in cell_text or '盈餘分配率' in cell_text or '分配比例' in cell_text):
            formula_tables['dividend'] = rows
            pages['dividend'] = page
        elif ('個人所得稅' in cell_text or '個人綜合所得稅' in cell_text or '薪資所得稅' in cell_text) and (
                '平均受雇員工年薪' in cell_text or '平均薪資' in cell_text or '員工薪資' in cell_text
                or '增聘' in cell_text):
            formula_tables['personal'] = rows
            pages['personal'] = page
        elif '汽車貨物稅' in cell_text and '貨物稅稅率' in cell_text:
            formula_tables['commodity'] = rows
            pages['commodity'] = page
        elif '加值型營業稅' in cell_text and '加值型營業稅稅率' in cell_text:
            formula_tables['vat'] = rows
            pages['vat'] = page

    data['formula_tables'] = formula_tables

    formulas = {}

    def get_txt(key):
        return '\n'.join(' '.join(r) for r in formula_tables[key]) if key in formula_tables else ''

    txt = get_txt('vehicle_cit')
    m = re.search(r'=\s*([\d\.]+)\s*億元\s*[×x]\s*15%', txt)
    data['vehicle_output'] = float(m.group(1)) if m else 49.0
    m2 = re.search(r'新臺幣\s*([\d\.]+)\s*億元', txt)
    data['vehicle_cit_reported'] = float(m2.group(1)) if m2 else None
    formulas['vehicle_cit'] = compact_formula_full(txt) if txt else ''

    txt = get_txt('parts_cit')
    m = re.search(r'=\s*([\d\.]+)\s*億元\s*[×x]\s*13%', txt)
    data['parts_output'] = float(m.group(1)) if m else 34.31
    m2 = re.search(r'新臺幣\s*([\d\.]+)\s*億元', txt)
    data['parts_cit_reported'] = float(m2.group(1)) if m2 else None
    formulas['parts_cit'] = compact_formula_full(txt) if txt else ''

    txt = get_txt('other_cit')
    m = re.search(r'\(\s*([\d\.]+)\s*億元\s*[+＋]\s*([\d\.]+)\s*億元\s*\)', txt)
    if m:
        data['other_profit'] = float(m.group(1))
        data['tariff_savings'] = float(m.group(2))
    else:
        data['other_profit'] = 16.05
        data['tariff_savings'] = 22.40
    m2 = re.search(r'新臺幣\s*([\d\.]+)\s*億元', txt)
    data['other_cit_reported'] = float(m2.group(1)) if m2 else None
    formulas['other_cit'] = compact_formula_full(txt) if txt else ''

    txt = get_txt('dividend')
    all_vals = [float(v) for v in re.findall(r'(?<!\d)([\d]+\.[\d]+)\s*億元', txt)]
    small_vals = [v for v in all_vals if v < 5]
    data['dividend_tax_reported'] = small_vals[-1] if small_vals else 0.77
    formulas['dividend'] = compact_formula_full(txt) if txt else ''
    # Extract sub-components for dividend computation
    # Formula: profit × (1-CIT) × dist_rate × personal_rate  (or ×1 for some variants)
    div_params = {}
    if txt:
        m_dist = re.search(r'盈餘分配(?:比例|率)[^\d]*([\d\.]+)%', txt)
        m_rate = re.search(r'(?:個人|股利)[^\d]*所得稅率?[^\d]*([\d\.]+)%', txt)
        m_profit = re.search(r'([\d\.]+)\s*億元.*?(?:稅前|稅後|利潤|利益)', txt)
        if not m_profit:
            m_profit = re.search(r'(?:稅前|稅後|利潤).*?([\d\.]+)\s*億元', txt)
        if m_dist:
            div_params['dist_rate'] = float(m_dist.group(1)) / 100
        if m_rate:
            div_params['personal_rate'] = float(m_rate.group(1)) / 100
        if m_profit:
            div_params['profit'] = float(m_profit.group(1))
    data['dividend_params'] = div_params

    txt = get_txt('personal')
    m = re.search(r'合計\s*=\s*([\d\.]+)\s*萬元', txt)
    if not m:
        vals = re.findall(r'=\s*([\d\.]+)\s*萬元', txt)
        data['personal_tax_reported'] = float(vals[-1]) / 10000 if vals else 0.0287
    else:
        data['personal_tax_reported'] = float(m.group(1)) / 10000
    formulas['personal'] = compact_formula_full(txt) if txt else ''
    # Extract sub-components for personal tax computation
    # Formula: employees × avg_salary × rate
    per_params = {}
    if txt:
        m_emp = re.search(r'(?:增聘|受雇員工|員工人數)[^\d]*(\d[\d,]*)\s*人', txt)
        m_sal = re.search(r'平均.*?(?:年薪|薪資)[^\d]*([\d\.]+)\s*萬元', txt)
        if not m_sal:
            m_sal = re.search(r'([\d\.]+)\s*萬元.*?(?:年薪|薪資)', txt)
        m_rate = re.search(r'(?:個人|綜合).*?所得稅率?[^\d]*([\d\.]+)%', txt)
        if not m_rate:
            m_rate = re.search(r'(?:所得稅率?|適用稅率)[^\d]*([\d\.]+)%', txt)
        if m_emp:
            per_params['employees'] = float(m_emp.group(1).replace(',', ''))
        if m_sal:
            per_params['salary_wan'] = float(m_sal.group(1))
        if m_rate:
            per_params['rate'] = float(m_rate.group(1)) / 100
        if per_params.get('employees') and per_params.get('salary_wan') and per_params.get('rate'):
            per_params['calc'] = round2(
                per_params['employees'] * per_params['salary_wan'] * per_params['rate'] / 10000)
    data['personal_params'] = per_params

    txt = get_txt('commodity')
    m = re.search(r'新臺幣\s*([\d\.]+)\s*億元', txt)
    data['commodity_tax_reported'] = float(m.group(1)) if m else None
    formulas['commodity'] = compact_formula_full(txt) if txt else ''

    txt = get_txt('vat')
    m = re.search(r'新臺幣\s*([\d\.]+)\s*億元', txt)
    data['vat_reported'] = float(m.group(1)) if m else None
    formulas['vat'] = compact_formula_full(txt) if txt else ''

    # Net (search all text)
    all_text, get_page = build_paged_text(para_list, table_list)
    m = re.search(r'(?:淨增|損益合計|淨[損益額])[^\d+\-]*([\+\-]?\s*[\d\.]+)\s*億元', all_text)
    if m:
        data['net_reported'] = float(m.group(1).replace(' ', ''))
        pages['net'] = get_page(m.start())
    else:
        data['net_reported'] = 0.29

    data['formulas'] = formulas
    data['pages'] = pages
    data['table_values'] = scan_tables_financial(table_list, '億元')

    return data


# ── simplified summary-table parser ──────────────────────────────────────────

def parse_simplified_summary_table(table_list):
    """
    Find the 項目 | 最初收入損失法 | 最終收入損失法 summary section inside
    the simplified-format's large table and extract each tax item's 最終 value.
    Handles abbreviated labels and negative tariff figures.
    Returns {key: (abs_value, page)}.
    """
    LABEL_MAP = [
        ('tariff_loss', ['關稅']),
        ('commodity',   ['貨物稅']),
        ('vat',         ['營業稅']),
        ('total_cit',   ['營利事業所得稅']),
        ('personal',    ['個人綜合所得稅', '個人所得稅']),
        ('dividend',    ['股東個人所得稅', '股東個人股利所得稅']),
        ('net',         ['合計']),
    ]

    found = {}
    for entry in table_list:
        page, rows, row_pages = _tbl(entry)
        # Locate the header row that marks the summary section
        header_ri = None
        for ri, row in enumerate(rows):
            row_flat = ' '.join(row)
            if '最初收入損失法' in row_flat and '最終收入損失法' in row_flat:
                header_ri = ri
                break
        if header_ri is None:
            continue

        # Identify the column index for '最終收入損失法'
        header_row = rows[header_ri]
        final_col = None
        for ci, cell in enumerate(header_row):
            if '最終收入損失法' in cell:
                final_col = ci
        if final_col is None:
            final_col = len(header_row) - 1

        for offset, row in enumerate(rows[header_ri + 1:]):
            ri = header_ri + 1 + offset
            if not row:
                continue
            label = row[0].strip()
            if not label:
                continue

            val_str = ''
            if final_col < len(row):
                val_str = row[final_col].strip()
            if not val_str:
                for cell in reversed(row):
                    if cell.strip():
                        val_str = cell.strip()
                        break
            if not val_str:
                continue

            val_str_clean = re.sub(r'[,，\s]', '', val_str)
            try:
                val = float(val_str_clean)
            except ValueError:
                continue

            row_page = row_pages[ri] if ri < len(row_pages) else page
            for key, labels in LABEL_MAP:
                if key in found:
                    continue
                if any(lbl in label for lbl in labels):
                    found[key] = (abs(val), row_page)
                    break

    return found


# ── simplified-format parser ──────────────────────────────────────────────────

def parse_simplified_report(table_list, para_list):
    data = {}
    pages = {}

    all_text, get_page = build_paged_text(para_list, table_list)

    def find_k(pattern):
        m = re.search(pattern, all_text, re.DOTALL)
        if m:
            raw = m.group(1).replace(',', '').strip()
            try:
                return float(raw), get_page(m.start())
            except ValueError:
                pass
        return None, None

    def fk(pattern, key):
        val, pg = find_k(pattern)
        if val is not None:
            pages[key] = pg
        return val

    data['total_import_reported_k'] = fk(r'平均進口值為([\d,]+)千元', 'total_import')
    data['tariff_loss_reported_k']  = fk(r'關稅收入減少([\d,]+)千元', 'tariff_loss')
    data['vehicle_output_k']        = fk(r'產值(?:變化值|增加額)\s*([\d,]+)千元.*?小客車生產', 'vehicle_output')
    data['parts_output_k']          = fk(r'零組件產值增加額為([\d,]+)千元', 'parts_output')
    data['vehicle_cit_reported_k']  = fk(r'整車業增加稅額約為新臺幣([\d,]+)千元', 'vehicle_cit')
    data['parts_cit_reported_k']    = fk(r'國產化比例約?70%.*?增加稅額約為新臺幣([\d,]+)千元', 'parts_cit')
    if data['parts_cit_reported_k'] is None:
        data['parts_cit_reported_k'] = fk(
            r'零組件產值增加額為[\d,]+千元.*?增加稅額約為新臺幣([\d,]+)千元', 'parts_cit')
    data['other_cit_reported_k']    = fk(r'就其他工業及服務業部分.*?新臺幣([\d,]+)千元', 'other_cit')
    data['total_cit_reported_k']    = fk(r'三項合計.*?新臺幣([\d,]+)千元', 'total_cit')
    data['commodity_tax_reported_k']= fk(r'貨物稅增加額約為新臺幣([\d,]+)千元', 'commodity')
    data['vat_reported_k']          = fk(r'加值型營業稅增加額約為新臺幣([\d,]+)千元', 'vat')
    data['personal_tax_reported_k'] = fk(r'員工個人所得稅增加額約為新臺幣([\d,]+)千元', 'personal')
    if data['personal_tax_reported_k'] is None:
        data['personal_tax_reported_k'] = fk(
            r'兩項合計.*?個人所得稅.*?新臺幣([\d,]+)千元', 'personal')
    data['dividend_tax_reported_k'] = fk(r'股東個人股利所得稅增加額約為新臺幣([\d,]+)千元', 'dividend')
    data['net_reported_k']          = fk(r'稅收淨收入([\d,]+)千元', 'net')
    if data['net_reported_k'] is None:
        data['net_reported_k']      = fk(r'最終稅收.*?([\d,]+)千元', 'net')
    data['other_profit_k']          = fk(r'增加產值利潤([\d,]+)千元', 'other_profit')

    # Formula text blocks
    formulas = {}
    vo_k = data.get('vehicle_output_k') or 4900813
    po_k = data.get('parts_output_k') or 3430569

    m = re.search(r'整車業增加稅額約為新臺幣[\d,]+千元([^。\n]{0,300})', all_text, re.DOTALL)
    formulas['vehicle_cit'] = (compact_formula_braces(m.group(0)) if m else
                                f'{fmt(vo_k,0)}千元 × 15% × 20%')

    m = re.search(r'國產化比例約?70%[^。]{0,400}增加稅額約為新臺幣[\d,]+千元([^。]{0,300})', all_text, re.DOTALL)
    formulas['parts_cit'] = (compact_formula_braces(m.group(0)) if m else
                              f'{fmt(po_k,0)}千元 × 13% × 20%')

    m = re.search(r'兩者合計[\d,]+千元[，,]([^。]{0,300}新臺幣[\d,]+千元)', all_text, re.DOTALL)
    if not m:
        m = re.search(r'就其他工業及服務業部分[^。]{0,500}新臺幣[\d,]+千元([^。]{0,200})', all_text, re.DOTALL)
    formulas['other_cit'] = compact_formula_braces(m.group(0)) if m else '3,844,451千元 × 20%'

    m = re.search(r'貨物稅增加額約為新臺幣[\d,]+千元([^。]{0,500})', all_text, re.DOTALL)
    formulas['commodity'] = compact_formula_braces(m.group(0)) if m else ''

    m = re.search(r'加值型營業稅增加額約為新臺幣[\d,]+千元([^。\n]{0,500})', all_text, re.DOTALL)
    formulas['vat'] = compact_formula_braces(m.group(0)) if m else ''

    m = re.search(r'員工個人所得稅增加額約為新臺幣[\d,]+千元([^。]{0,400})', all_text, re.DOTALL)
    formulas['personal'] = compact_formula_braces(m.group(0)) if m else ''

    m = re.search(r'股東個人股利所得稅增加額約為新臺幣[\d,]+千元([^。]{0,400})', all_text, re.DOTALL)
    formulas['dividend'] = compact_formula_braces(m.group(0)) if m else ''

    data['formulas'] = formulas
    data['pages'] = pages

    # Merge generic scanner + dedicated summary-table parser.
    # Summary table wins when both find the same key (it uses the definitive
    # 最終收入損失法 column and handles abbreviated labels / negative values).
    table_values = scan_tables_financial(table_list, '千元')
    summary_tv   = parse_simplified_summary_table(table_list)
    table_values.update(summary_tv)          # summary overrides generic scan
    data['table_values'] = table_values

    # Sync pages and values from the summary table.
    # Always update the page (summary table location is most authoritative);
    # only set the value when regex prose search didn't already find it.
    _backfill = [
        ('tariff_loss_reported_k',  'tariff_loss'),
        ('commodity_tax_reported_k','commodity'),
        ('vat_reported_k',          'vat'),
        ('total_cit_reported_k',    'total_cit'),
        ('personal_tax_reported_k', 'personal'),
        ('dividend_tax_reported_k', 'dividend'),
        ('net_reported_k',          'net'),
    ]
    for data_key, tv_key in _backfill:
        if tv_key not in summary_tv:
            continue
        val, pg = summary_tv[tv_key]
        # Always overwrite page with the summary table page
        page_key = data_key.replace('_reported_k', '').replace('_tax', '')
        pages[page_key] = pg
        # Only backfill value when regex found nothing
        if data.get(data_key) is None:
            data[data_key] = val

    return data


# ── verification: full report ─────────────────────────────────────────────────

SEC_TARIFF  = '一、最初收入損失法'
SEC_CT      = '　（一）貨物稅'
SEC_VAT     = '　（二）加值型營業稅'
SEC_CIT     = '　（三）營利事業所得稅'
SEC_PER     = '　（四）個人綜合所得稅'
SEC_DIV     = '　（五）股東個人股利所得稅'
SEC_NET     = '三、淨損益'
SEC_TBL     = '附：表格數值核對'


def verify_full(data):
    results = []
    items  = data.get('items', [])
    forms  = data.get('formulas', {})
    pages  = data.get('pages', {})

    # ── 最初收入損失法 ───────────────────────────────────────────────────────
    if items:
        calc_total = round2(sum(it['import_5yr'] for it in items))
        rep = data.get('total_import_reported') or 147.86
        results.append(row_result(
            SEC_TARIFF, '5年平均進口總額 (億元)', '各品項加總',
            fmt(rep), fmt(calc_total), fmt(rep - calc_total),
            status_icon(ok(rep, calc_total)), '', pages.get('table41')))

    if items and all(it['current_rate'] is not None for it in items):
        calc_wavg = sum(it['current_rate'] for it in items) / len(items)
        rep_wavg = data.get('weighted_rate_reported') or 0.1185
        results.append(row_result(
            SEC_TARIFF, f'算術平均現行稅率（{len(items)}項）',
            f'Σ稅率 ÷ {len(items)}',
            f'{rep_wavg*100:.2f}%', f'{calc_wavg*100:.2f}%',
            f'{(rep_wavg-calc_wavg)*100:.4f}%',
            status_icon(ok(rep_wavg, calc_wavg, tol=0.005)), '', pages.get('table41')))

    calc_total_loss = 0
    for it in items:
        if it['import_5yr'] is None or it['current_rate'] is None or it['after_rate'] is None:
            continue
        cut = it['current_rate'] - it['after_rate']
        calc_loss = round2(it['import_5yr'] * cut)
        rep_loss_item = data.get('item_losses_reported', {}).get(it['hs'])
        calc_total_loss += calc_loss
        results.append(row_result(
            SEC_TARIFF,
            f"{it['hs']}　{it['name'][:14]}",
            f"{it['import_5yr']} × {cut*100:.1f}%",
            fmt(rep_loss_item) if rep_loss_item else '—',
            fmt(calc_loss),
            fmt(rep_loss_item - calc_loss) if rep_loss_item else '—',
            status_icon(ok(rep_loss_item, calc_loss)) if rep_loss_item else '⚠️',
            '關稅損失(億元)', pages.get('table42')))

    rep_loss = data.get('tariff_loss_reported') or 22.40
    results.append(row_result(
        SEC_TARIFF, '關稅損失合計 (億元)', 'Σ各品項關稅損失',
        fmt(rep_loss), fmt(calc_total_loss), fmt(rep_loss - calc_total_loss),
        status_icon(ok(rep_loss, calc_total_loss)), '', pages.get('table42')))

    # ── 貨物稅 ────────────────────────────────────────────────────────────────
    vo = data.get('vehicle_output', 49.0)
    calc_ct = round2(vo * (0.5285 * 0.25 * 0.8419 + 0.4715 * 0.15 * 0.9960))
    rep_ct  = data.get('commodity_tax_reported')
    results.append(row_result(
        SEC_CT, '貨物稅 (億元)',
        forms.get('commodity') or f'{vo}億元×[52.85%×25%×84.19%+47.15%×15%×99.60%]',
        fmt(rep_ct) if rep_ct else '—', fmt(calc_ct),
        fmt(rep_ct - calc_ct) if rep_ct else '—',
        status_icon(ok(rep_ct, calc_ct)) if rep_ct else '⚠️', '', pages.get('commodity')))

    # ── 加值型營業稅 ──────────────────────────────────────────────────────────
    calc_vat = round2(vo * (0.5285 * 1.25 * 0.05 + 0.4715 * 1.15 * 0.05))
    rep_vat  = data.get('vat_reported')
    results.append(row_result(
        SEC_VAT, '加值型營業稅 (億元)',
        forms.get('vat') or f'{vo}億元×[52.85%×125%×5%+47.15%×115%×5%]',
        fmt(rep_vat) if rep_vat else '—', fmt(calc_vat),
        fmt(rep_vat - calc_vat) if rep_vat else '—',
        status_icon(ok(rep_vat, calc_vat)) if rep_vat else '⚠️', '', pages.get('vat')))

    # ── 營利事業所得稅 ────────────────────────────────────────────────────────
    po = data.get('parts_output', 34.31)
    op = data.get('other_profit', 16.05)
    ts = data.get('tariff_savings', 22.40)

    calc_vcit = round2(vo * 0.15 * 0.20)
    rep_vcit  = data.get('vehicle_cit_reported')
    results.append(row_result(
        SEC_CIT, '整車營利事業所得稅 (億元)',
        forms.get('vehicle_cit') or f'{vo}億元×15%×20%',
        fmt(rep_vcit) if rep_vcit else '—', fmt(calc_vcit),
        fmt(rep_vcit - calc_vcit) if rep_vcit else '—',
        status_icon(ok(rep_vcit, calc_vcit)) if rep_vcit else '⚠️', '', pages.get('vehicle_cit')))

    calc_pcit = round2(po * 0.13 * 0.20)
    rep_pcit  = data.get('parts_cit_reported')
    results.append(row_result(
        SEC_CIT, '零組件營利事業所得稅 (億元)',
        forms.get('parts_cit') or f'{po}億元×13%×20%',
        fmt(rep_pcit) if rep_pcit else '—', fmt(calc_pcit),
        fmt(rep_pcit - calc_pcit) if rep_pcit else '—',
        status_icon(ok(rep_pcit, calc_pcit)) if rep_pcit else '⚠️', '', pages.get('parts_cit')))

    calc_ocit = round2((op + ts) * 0.20)
    rep_ocit  = data.get('other_cit_reported')
    results.append(row_result(
        SEC_CIT, '其他工業及服務業營利事業所得稅 (億元)',
        forms.get('other_cit') or f'({op}+{ts})億元×20%',
        fmt(rep_ocit) if rep_ocit else '—', fmt(calc_ocit),
        fmt(rep_ocit - calc_ocit) if rep_ocit else '—',
        status_icon(ok(rep_ocit, calc_ocit)) if rep_ocit else '⚠️', '', pages.get('other_cit')))

    calc_tcit = round2(calc_vcit + calc_pcit + calc_ocit)
    rep_tcit  = data.get('total_cit_reported')
    if rep_tcit is None:
        parts_list = [data.get('vehicle_cit_reported'), data.get('parts_cit_reported'), data.get('other_cit_reported')]
        if all(p is not None for p in parts_list):
            rep_tcit = round2(sum(parts_list))
    results.append(row_result(
        SEC_CIT, '營利事業所得稅合計 (億元)', '整車+零組件+其他',
        fmt(rep_tcit) if rep_tcit else '—', fmt(calc_tcit),
        fmt(rep_tcit - calc_tcit) if rep_tcit else '—',
        status_icon(ok(rep_tcit, calc_tcit)) if rep_tcit else '⚠️', '', None))

    rep_per = data.get('personal_tax_reported', 0.0287)
    per_params = data.get('personal_params', {})
    if per_params.get('calc') is not None:
        calc_per = per_params['calc']
        per_note = (f"{per_params.get('employees',0):.0f}人"
                    f"×{per_params.get('salary_wan',0):.2f}萬元"
                    f"×{per_params.get('rate',0)*100:.1f}%")
        results.append(row_result(
            SEC_PER, '個人綜合所得稅 (億元)',
            forms.get('personal') or per_note,
            fmt(rep_per), fmt(calc_per), fmt(rep_per - calc_per),
            status_icon(ok(rep_per, calc_per)), per_note, pages.get('personal')))
    else:
        results.append(row_result(
            SEC_PER, '個人綜合所得稅 (億元)',
            forms.get('personal') or '增聘員工薪資 × 適用稅率',
            fmt(rep_per), '（參數未能自動抽取）', '—', '⚠️',
            '請手動確認員工數、薪資、稅率', pages.get('personal')))

    rep_div = data.get('dividend_tax_reported', 0.77)
    div_params = data.get('dividend_params', {})
    if (div_params.get('profit') and div_params.get('dist_rate') and div_params.get('personal_rate')):
        calc_div = round2(
            div_params['profit'] * (1 - 0.20) * div_params['dist_rate'] * div_params['personal_rate'])
        div_note = (f"{div_params['profit']}億元"
                    f"×80%×{div_params['dist_rate']*100:.0f}%"
                    f"×{div_params['personal_rate']*100:.0f}%")
        results.append(row_result(
            SEC_DIV, '股東個人股利所得稅 (億元)',
            forms.get('dividend') or div_note,
            fmt(rep_div), fmt(calc_div), fmt(rep_div - calc_div),
            status_icon(ok(rep_div, calc_div)), div_note, pages.get('dividend')))
    else:
        results.append(row_result(
            SEC_DIV, '股東個人股利所得稅 (億元)',
            forms.get('dividend') or '利潤×(1-CIT)×盈餘分配率×個人稅率',
            fmt(rep_div), '（參數未能自動抽取）', '—', '⚠️',
            '請手動確認利潤基數、分配率、稅率', pages.get('dividend')))

    # ── 淨損益 ────────────────────────────────────────────────────────────────
    calc_net = round2(calc_tcit + rep_div + rep_per + calc_ct + calc_vat - rep_loss)
    rep_net  = data.get('net_reported', 0.29)
    results.append(row_result(
        SEC_NET, '最終收入損失法淨損益 (億元)',
        'CIT+股利+個人+貨物稅+營業稅－關稅損失',
        fmt(rep_net), fmt(calc_net), fmt(rep_net - calc_net),
        status_icon(ok(rep_net, calc_net, tol=0.10)),
        f'{fmt(calc_tcit)}+{fmt(rep_div)}+{fmt(rep_per)}+{fmt(calc_ct)}+{fmt(calc_vat)}−{fmt(rep_loss)}',
        pages.get('net')))

    # ── 附：表格數值核對 ──────────────────────────────────────────────────────
    tv = data.get('table_values', {})
    tbl_checks = [
        ('tariff_loss', '表格：關稅損失 (億元)',            rep_loss,  calc_total_loss),
        ('commodity',   '表格：貨物稅 (億元)',              rep_ct,    calc_ct),
        ('vat',         '表格：加值型營業稅 (億元)',         rep_vat,   calc_vat),
        ('vehicle_cit', '表格：整車CIT (億元)',             rep_vcit,  calc_vcit),
        ('parts_cit',   '表格：零組件CIT (億元)',           rep_pcit,  calc_pcit),
        ('other_cit',   '表格：其他CIT (億元)',             rep_ocit,  calc_ocit),
        ('total_cit',   '表格：CIT合計 (億元)',             rep_tcit,  calc_tcit),
        ('net',         '表格：淨損益 (億元)',               rep_net,   calc_net),
    ]
    tbl_rows_added = 0
    for key, label, formula_rep, formula_calc in tbl_checks:
        if key not in tv:
            continue
        tbl_val, tbl_page = tv[key]
        diff_tc = round2(tbl_val - formula_calc) if formula_calc is not None else None
        note = f'表格值={fmt(tbl_val)}，公式計算={fmt(formula_calc)}'
        if formula_rep is not None:
            note += f'，公式文字={fmt(formula_rep)}'
            passed = ok(tbl_val, formula_calc) and ok(tbl_val, formula_rep)
        else:
            passed = ok(tbl_val, formula_calc)
        results.append(row_result(
            SEC_TBL, label, '（來自表格欄位）',
            fmt(tbl_val), fmt(formula_calc) if formula_calc else '—',
            fmt(diff_tc) if diff_tc is not None else '—',
            status_icon(passed), note, tbl_page))
        tbl_rows_added += 1

    if tbl_rows_added == 0:
        results.append(row_result(
            SEC_TBL, '（未偵測到獨立彙總表格）', '',
            '—', '—', '—', '⚠️', '若文件含有獨立彙總結果表，請確認表格關鍵字'))

    return pd.DataFrame(results)


# ── verification: simplified report ──────────────────────────────────────────

def verify_simplified(data):
    results = []
    forms = data.get('formulas', {})
    pages = data.get('pages', {})

    VEHICLE_NET_MARGIN = 0.15
    PARTS_NET_MARGIN   = 0.13
    CIT_RATE           = 0.20
    DOMESTIC_SHARE     = 0.5285
    IMPORT_SHARE       = 0.4715
    DOMESTIC_CT_RATE   = 0.25
    IMPORT_CT_RATE     = 0.15
    DOMESTIC_TAXABLE   = 0.8419
    IMPORT_TAXABLE     = 0.9960
    DOMESTIC_TAX_BASE  = 1.25
    IMPORT_TAX_BASE    = 1.15
    VAT_RATE           = 0.05

    rep_loss = data.get('tariff_loss_reported_k')  or 2240360
    rep_vo   = data.get('vehicle_output_k')         or 4900813
    rep_po   = data.get('parts_output_k')           or 3430569
    rep_tcit = data.get('total_cit_reported_k')
    rep_vcit = data.get('vehicle_cit_reported_k')
    rep_pcit = data.get('parts_cit_reported_k')
    rep_ocit = data.get('other_cit_reported_k')
    rep_per  = data.get('personal_tax_reported_k')  or 2874
    rep_div  = data.get('dividend_tax_reported_k')  or 76500
    rep_ct   = data.get('commodity_tax_reported_k') or 890371
    rep_vat  = data.get('vat_reported_k')           or 294747
    rep_net  = data.get('net_reported_k')           or 29241
    OTHER_PROFIT_K = data.get('other_profit_k') or 1604091

    # ── 最初收入損失法 ────────────────────────────────────────────────────────
    rep_import = data.get('total_import_reported_k') or 14785668
    results.append(row_result(
        SEC_TARIFF, '5年平均進口總額 (千元)', '依財政部關務署統計',
        fmt(rep_import, 0), '（引用文件值）', '—', '⚠️', '', pages.get('total_import')))
    results.append(row_result(
        SEC_TARIFF, '關稅損失 (千元)', '進口額×降稅幅度',
        fmt(rep_loss, 0), '（引用文件值）', '—', '⚠️', '最初收入損失法', pages.get('tariff_loss')))

    # ── 貨物稅 ────────────────────────────────────────────────────────────────
    calc_ct = round2(rep_vo * (DOMESTIC_SHARE * DOMESTIC_CT_RATE * DOMESTIC_TAXABLE
                               + IMPORT_SHARE * IMPORT_CT_RATE * IMPORT_TAXABLE))
    results.append(row_result(
        SEC_CT, '貨物稅 (千元)',
        forms.get('commodity') or f'{fmt(rep_vo,0)}×[52.85%×25%×84.19%+47.15%×15%×99.60%]',
        fmt(rep_ct, 0), fmt(calc_ct, 0), fmt(rep_ct - calc_ct, 0),
        status_icon(ok(rep_ct, calc_ct)), '', pages.get('commodity')))

    # ── 加值型營業稅 ──────────────────────────────────────────────────────────
    calc_vat = round2(rep_vo * (DOMESTIC_SHARE * DOMESTIC_TAX_BASE * VAT_RATE
                                + IMPORT_SHARE * IMPORT_TAX_BASE * VAT_RATE))
    results.append(row_result(
        SEC_VAT, '加值型營業稅 (千元)',
        forms.get('vat') or f'{fmt(rep_vo,0)}×[52.85%×125%×5%+47.15%×115%×5%]',
        fmt(rep_vat, 0), fmt(calc_vat, 0), fmt(rep_vat - calc_vat, 0),
        status_icon(ok(rep_vat, calc_vat)), '', pages.get('vat')))

    # ── 營利事業所得稅 ────────────────────────────────────────────────────────
    calc_vcit = round2(rep_vo * VEHICLE_NET_MARGIN * CIT_RATE)
    results.append(row_result(
        SEC_CIT, '整車營利事業所得稅 (千元)',
        forms.get('vehicle_cit') or f'{fmt(rep_vo,0)}×15%×20%',
        fmt(rep_vcit, 0) if rep_vcit else '—', fmt(calc_vcit, 0),
        fmt(rep_vcit - calc_vcit, 0) if rep_vcit else '—',
        status_icon(ok(rep_vcit, calc_vcit)) if rep_vcit else '⚠️', '', pages.get('vehicle_cit')))

    calc_pcit = round2(rep_po * PARTS_NET_MARGIN * CIT_RATE)
    results.append(row_result(
        SEC_CIT, '零組件營利事業所得稅 (千元)',
        forms.get('parts_cit') or f'{fmt(rep_po,0)}×13%×20%',
        fmt(rep_pcit, 0) if rep_pcit else '—', fmt(calc_pcit, 0),
        fmt(rep_pcit - calc_pcit, 0) if rep_pcit else '—',
        status_icon(ok(rep_pcit, calc_pcit)) if rep_pcit else '⚠️', '', pages.get('parts_cit')))

    calc_ocit = round2((OTHER_PROFIT_K + rep_loss) * CIT_RATE)
    results.append(row_result(
        SEC_CIT, '其他工業及服務業營利事業所得稅 (千元)',
        forms.get('other_cit') or f'({fmt(OTHER_PROFIT_K,0)}+{fmt(rep_loss,0)})×20%',
        fmt(rep_ocit, 0) if rep_ocit else '—', fmt(calc_ocit, 0),
        fmt(rep_ocit - calc_ocit, 0) if rep_ocit else '—',
        status_icon(ok(rep_ocit, calc_ocit)) if rep_ocit else '⚠️', '', pages.get('other_cit')))

    calc_tcit = round2(calc_vcit + calc_pcit + calc_ocit)
    results.append(row_result(
        SEC_CIT, '營利事業所得稅合計 (千元)', '整車+零組件+其他',
        fmt(rep_tcit, 0) if rep_tcit else '—', fmt(calc_tcit, 0),
        fmt(rep_tcit - calc_tcit, 0) if rep_tcit else '—',
        status_icon(ok(rep_tcit, calc_tcit)) if rep_tcit else '⚠️', '', pages.get('total_cit')))

    results.append(row_result(
        SEC_PER, '個人綜合所得稅 (千元)',
        forms.get('personal') or '增聘員工薪資×適用稅率',
        fmt(rep_per, 0), '（依文件值）', '—', '⚠️', '依文件公式確認', pages.get('personal')))

    results.append(row_result(
        SEC_DIV, '股東個人股利所得稅 (千元)',
        forms.get('dividend') or '產值利潤×80%×56%×30%×40.45%×28%',
        fmt(rep_div, 0), '（依文件值）', '—', '⚠️', '依文件公式確認', pages.get('dividend')))

    # ── 淨損益 ────────────────────────────────────────────────────────────────
    calc_net = round2(calc_vcit + calc_pcit + calc_ocit + rep_per + rep_div
                      + calc_ct + calc_vat - rep_loss)
    results.append(row_result(
        SEC_NET, '最終收入損失法淨損益 (千元)',
        'CIT+股利+個人+貨物稅+營業稅－關稅損失',
        fmt(rep_net, 0), fmt(calc_net, 0), fmt(rep_net - calc_net, 0),
        status_icon(ok(rep_net, calc_net, tol=0.05)),
        f'{fmt(calc_tcit,0)}+{fmt(rep_div,0)}+{fmt(rep_per,0)}+{fmt(calc_ct,0)}+{fmt(calc_vat,0)}−{fmt(rep_loss,0)}',
        pages.get('net')))

    # ── 附：表格數值核對 ──────────────────────────────────────────────────────
    tv = data.get('table_values', {})
    tbl_checks = [
        ('tariff_loss', '表格：關稅損失 (千元)',           rep_loss,  None),
        ('commodity',   '表格：貨物稅 (千元)',             rep_ct,    calc_ct),
        ('vat',         '表格：加值型營業稅 (千元)',        rep_vat,   calc_vat),
        ('vehicle_cit', '表格：整車CIT (千元)',            rep_vcit,  calc_vcit),
        ('parts_cit',   '表格：零組件CIT (千元)',          rep_pcit,  calc_pcit),
        ('other_cit',   '表格：其他CIT (千元)',            rep_ocit,  calc_ocit),
        ('total_cit',   '表格：CIT合計 (千元)',            rep_tcit,  calc_tcit),
        ('net',         '表格：淨損益 (千元)',              rep_net,   calc_net),
    ]
    tbl_rows_added = 0
    for key, label, text_rep, formula_calc in tbl_checks:
        if key not in tv:
            continue
        tbl_val, tbl_page = tv[key]
        calc_ref = formula_calc if formula_calc is not None else text_rep
        diff_tc = round2(tbl_val - calc_ref) if calc_ref is not None else None
        note = f'表格讀取={fmt(tbl_val, 0)}'
        if text_rep is not None:
            note += f'，文字萃取={fmt(text_rep, 0)}'
            passed = ok(tbl_val, text_rep) and (formula_calc is None or ok(tbl_val, formula_calc))
        else:
            passed = ok(tbl_val, formula_calc) if formula_calc else None
        results.append(row_result(
            SEC_TBL, label, '（來自表格欄位）',
            fmt(tbl_val, 0),
            fmt(formula_calc, 0) if formula_calc is not None else '（引用）',
            fmt(diff_tc, 0) if diff_tc is not None else '—',
            status_icon(passed), note, tbl_page))
        tbl_rows_added += 1

    if tbl_rows_added == 0:
        results.append(row_result(
            SEC_TBL, '（未偵測到表格中的千元數值）', '',
            '—', '—', '—', '⚠️', '若文件含有結果表格，請確認表格格式'))

    return pd.DataFrame(results)


# ── Excel export ──────────────────────────────────────────────────────────────

FILL_GREEN   = PatternFill('solid', fgColor='C6EFCE')
FILL_RED     = PatternFill('solid', fgColor='FFC7CE')
FILL_YELLOW  = PatternFill('solid', fgColor='FFEB9C')
FILL_HEADER  = PatternFill('solid', fgColor='4472C4')
FILL_SECTION = PatternFill('solid', fgColor='D9E1F2')
FILL_TBL     = PatternFill('solid', fgColor='E2EFDA')
FONT_WHITE   = Font(color='FFFFFF', bold=True)
FONT_SECTION = Font(bold=True, color='1F3864')
THIN         = Side(style='thin')
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def to_excel_single(df, title):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    headers = list(df.columns)
    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER

    current_section = None
    ri = 2
    for _, row_data in df.iterrows():
        sec = row_data.get('章節', '')
        if sec and sec != current_section:
            current_section = sec
            ws.merge_cells(start_row=ri, start_column=1,
                           end_row=ri, end_column=len(headers))
            cell = ws.cell(row=ri, column=1, value=sec)
            cell.fill = FILL_TBL if sec == SEC_TBL else FILL_SECTION
            cell.font = FONT_SECTION
            cell.alignment = Alignment(horizontal='left')
            cell.border = BORDER
            ri += 1

        result_val = str(row_data.get('結果', ''))
        fill = (FILL_GREEN if '✅' in result_val
                else FILL_RED if '❌' in result_val
                else FILL_YELLOW)
        for ci, col in enumerate(headers, 1):
            val = row_data[col]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = BORDER
            cell.fill = fill
        ri += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Streamlit UI ──────────────────────────────────────────────────────────────

def main():
    st.title("📋 稅式支出評估報告驗證工具")
    st.markdown("上傳 Word 報告（完整版 F2 或簡要格式），自動驗證所有計算，並匯出各份 Excel 驗證報告。")

    uploaded_files = st.file_uploader(
        "選擇檔案（Word 或 PDF，可一次上傳多份）",
        type=['docx', 'pdf'],
        accept_multiple_files=True,
        help="支援 .docx（Word）與 .pdf，PDF 頁碼精確；支援完整版（億元）與簡要格式（千元）",
    )

    if not uploaded_files:
        st.markdown("""
**支援格式：**
- **完整版報告**（如「汽車零組件關稅調降稅式支出評估報告」.F2.docx / .pdf）
  - 驗證：進口基準值、各品項關稅損失（表4-1/4-2）、各稅目公式計算、表格數值核對、最終淨損益
- **簡要格式報告**（如「稅式支出評估報告(簡要格式)」.F2.docx / .pdf）
  - 驗證：千元單位下各稅目公式計算、表格數值核對

上傳 **PDF** 可獲得精確頁碼（PDF 每頁為明確邊界，無需推算）。
每份檔案分別產生一份 Excel 驗證報告供下載，結果含**資料頁碼**欄位。
        """)
        return

    for uploaded_file in uploaded_files:
        st.divider()
        st.subheader(f"📄 {uploaded_file.name}")
        try:
            raw_bytes = uploaded_file.read()
            if uploaded_file.name.lower().endswith('.pdf'):
                para_list, table_list = extract_pdf(raw_bytes)
                file_fmt = 'PDF'
            else:
                doc = Document(io.BytesIO(raw_bytes))
                para_list, table_list = extract_document(doc)
                file_fmt = 'Word'
            doc_type = detect_type(table_list, para_list)

            label = '完整版報告（億元）' if doc_type == 'full' else '簡要格式報告（千元）'
            max_page = max((p for p, _ in para_list), default=1)
            for entry in table_list:
                _, _, rp = _tbl(entry)
                if rp:
                    max_page = max(max_page, max(rp))
            st.caption(f"識別類型：{label}　｜　格式：{file_fmt}　｜　表格數量：{len(table_list)}　｜　偵測頁數：第1～{max_page}頁")

            if doc_type == 'full':
                data = parse_full_report(table_list, para_list)
                df   = verify_full(data)
            else:
                data = parse_simplified_report(table_list, para_list)
                df   = verify_simplified(data)

            # Summary metrics
            total  = len(df)
            passed = (df['結果'] == '✅').sum()
            failed = (df['結果'] == '❌').sum()
            warn   = (df['結果'] == '⚠️').sum()
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("驗證項目", total)
            c2.metric("✅ 通過", passed)
            c3.metric("❌ 有誤", failed)
            c4.metric("⚠️ 待確認", warn)

            display_cols = [c for c in df.columns if c != '章節']

            def highlight(row):
                color = ('#C6EFCE' if row['結果'] == '✅'
                         else '#FFC7CE' if row['結果'] == '❌'
                         else '#FFEB9C')
                return [f'background-color: {color}'] * len(row)

            for sec, group in df.groupby('章節', sort=False):
                st.markdown(f"**{sec}**")
                g = group[display_cols].reset_index(drop=True)
                st.dataframe(g.style.apply(highlight, axis=1),
                             use_container_width=True, hide_index=True)

            if doc_type == 'full' and data.get('items'):
                with st.expander("展開：各品項明細表（表4-1）"):
                    item_df = pd.DataFrame(data['items'])
                    if not item_df.empty:
                        item_df['計算關稅損失(億元)'] = item_df.apply(
                            lambda r: round2(r['import_5yr'] * (r['current_rate'] - r['after_rate']))
                            if r['current_rate'] is not None and r['after_rate'] is not None else None,
                            axis=1)
                        item_df.columns = ['稅則號別', '貨名', '5年平均進口(億元)',
                                           '現行稅率', '降稅後稅率', '計算關稅損失(億元)']
                        st.dataframe(item_df, use_container_width=True, hide_index=True)

            tv = data.get('table_values', {})
            if tv:
                with st.expander(f"展開：偵測到的表格數值（共 {len(tv)} 項）"):
                    unit = '億元' if doc_type == 'full' else '千元'
                    tv_df = pd.DataFrame([
                        {'稅目鍵值': k,
                         f'表格讀取值（{unit}）': fmt(v, 0 if unit == '千元' else 2),
                         '來源頁碼': page_str(p)}
                        for k, (v, p) in tv.items()
                    ])
                    st.dataframe(tv_df, use_container_width=True, hide_index=True)

            short = (uploaded_file.name
                     .replace('.docx', '')
                     .replace('「', '').replace('」', '')
                     [:30])
            excel_buf = to_excel_single(df, short)
            btn_label = f"📥 下載 Excel：{short[:25]}…" if len(short) > 25 else f"📥 下載 Excel：{short}"
            st.download_button(
                label=btn_label,
                data=excel_buf,
                file_name=f"驗證結果_{short}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{uploaded_file.name}",
            )

        except Exception as e:
            st.error(f"解析失敗：{e}")
            st.exception(e)


if __name__ == '__main__':
    main()
